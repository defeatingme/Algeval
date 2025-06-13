import pandas as pd

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

from PySide6.QtWidgets import QFileDialog, QMessageBox
from styles import mboxStyle 


def export_data(ui_table):
    msg = mboxStyle()
    msg.setWindowTitle("Export records as...")
    msg.setText("Select file format:")
    msg.setIcon(QMessageBox.Question)

    # Remove default buttons
    msg.setStandardButtons(QMessageBox.NoButton)  # Ensures no extra buttons

    # Add only the required buttons
    pdf_button = msg.addButton("PDF", QMessageBox.AcceptRole)
    xlsx_button = msg.addButton("XLSX", QMessageBox.AcceptRole)
    msg.addButton("Cancel", QMessageBox.RejectRole)

    msg.exec()

    # Determine selected format
    if msg.clickedButton() == pdf_button:
        file_type = "pdf"
    elif msg.clickedButton() == xlsx_button:
        file_type = "xlsx"
    else:
        return  # Cancel export

    # Open file save dialog
    file_path, _ = QFileDialog.getSaveFileName(None, "Save File", "", f"{file_type.upper()} Files (*.{file_type})")
    if not file_path:
        return  # User canceled

    # Extract table data
    data = []
    for row in range(ui_table.rowCount()):
        row_data = []
        for col in range(1, ui_table.columnCount()):
            item = ui_table.item(row, col)
            row_data.append(item.text() if item else "")
        data.append(row_data)

    headers = ["Name", "Solution", "Final Ans.", "Overall"]  # Adjust headers as needed

    if file_type == "xlsx":
        save_xlsx(file_path, headers, data)
    elif file_type == "pdf":
        save_pdf(file_path, headers, data)

    mboxStyle.information(None, "Export Successful", f"Data successfully exported as {file_type.upper()}!")


def save_xlsx(file_path, headers, data):
    # Saves the table data as an Excel (.xlsx) file with auto-resized columns.
    wb = Workbook()
    ws = wb.active
    ws.append(headers)  # Write headers

    for row in data:
        ws.append([f"'{cell}" if isinstance(cell, str) and ("=" in cell or "/" in cell) else cell for cell in row])

    # Auto-adjust column width
    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2  # Add padding

    wb.save(file_path)


def save_csv(file_path, headers, data):
    # Save data to a CSV file.
    df = pd.DataFrame(data, columns=headers)
    df.to_csv(file_path, index=False, encoding="utf-8")


def save_pdf(file_path, headers, data):
    # Export data to a properly formatted PDF with adjustable columns
    
    doc = SimpleDocTemplate(file_path, pagesize=letter)
    
    # Get page width dynamically
    page_width, _ = letter  
    table_margin = 40  # Left & right margin to prevent cutoff
    
    # Define column width percentages (should total to 1.0 or 100%)
    col_ratios = [0.35, 0.25, 0.20, 0.20]  # Adjust as needed
    total_width = page_width - (table_margin * 2)  # Account for margins
    col_widths = [total_width * ratio for ratio in col_ratios]

    # Create table data (headers + rows)
    table_data = [headers] + data

    # Create table with dynamically adjusted column widths
    table = Table(table_data, colWidths=col_widths)

    # Apply table styles for better readability
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Header background
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),  # Header text color
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Center align all cells
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Bold headers
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),  # Extra padding for headers
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),  # White background for rows
        ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Add black grid lines
    ]))

    # Build the document with the table
    doc.build([table])